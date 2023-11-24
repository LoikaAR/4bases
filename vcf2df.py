import os
import csv
import pandas as pd
import mysql.connector
from mysql.connector import Error
from openpyxl import load_workbook

connection = None       # prepare connection
DP_THRESHOLD = 20       # depth threshold

db_config = {
    'host': 'localhost',
    'database': '4evar_test',
    'user': 'root',
    'password': 'PassMass123!'
}

def vcf_scraper(file_path):
    file_name = file_path.split('/')[-1]
    try:
        connection = mysql.connector.connect(**db_config)

        if connection and connection.is_connected():
            db_info = connection.get_server_info()
            print("Connected to server ", db_info)
            cursor = connection.cursor()
            cursor.execute("SELECT DATABASE();")
            db_name = cursor.fetchone()[0]
            print("Connected to database", db_name)

            with open(file_path, 'r') as file_:
                for i, line in enumerate(file_):
                    if line.startswith('#CHROM'):
                        header = line.strip('#').strip().split('\t')
                        break

            # read the vcf file
            data = pd.read_csv(file_path, sep ='\t', skiprows=i, header=None)
            pd.set_option('display.max_rows', data.shape[0]+1)
            pd.set_option('display.max_columns', 20)

            data.columns = header
            target_data = data[['CHROM', 'POS', 'REF', 'ALT']]
            n_rows = target_data.shape[0]
            n_cols = target_data.shape[1]

            # prepare the tsv file for existing variants
            out_file = open("existing.tsv", "a")
            tsv_writer = csv.writer(out_file, delimiter='\t')
            col_names = ["CHROM", "POS", "REF", "ALT", "count"]
            tsv_writer.writerow(col_names)

            # check every row in vcf file
            for row in range(1, n_rows):
                extracted_info = {"CHROM": None, "REF": None, "POS": None, "ALT": None}
                query_info = {"VAR_STRING": ""}

                for col in range(0, n_cols):
                    query_info['VAR_STRING'] += str(target_data.iloc[row, col])

                    # separate pieces of data for lookup in excel files 
                    if target_data.iloc[0,col] == "#CHROM":
                        extracted_info["CHROM"] = target_data.iloc[row, col]
                    elif target_data.iloc[0,col] == "REF":
                        extracted_info["REF"] = target_data.iloc[row, col]
                    elif target_data.iloc[0,col] == "POS":
                        extracted_info["POS"] = target_data.iloc[row, col]
                    elif target_data.iloc[0,col] == "ALT":
                        extracted_info["ALT"] = target_data.iloc[row, col]
                
                vs = query_info["VAR_STRING"]
                print("extracted var string:", vs)

                query_varstring = ("SELECT * FROM variant WHERE VAR_STRING = %(VAR_STRING)s") #  ask about concatenation

                cursor.execute(query_varstring, query_info)
                res = cursor.fetchall()

                # if variant is not found in the db
                if res == []:
                    print("Variant from vcf file not found in db - scraping excel file and inserting into db")

                    # ASSUMPTION: the excel file is in the same folder as the .vcf file
                    new_path = file_path.replace(file_name, '')     # strip path of the vcf file
                    folder = os.listdir(new_path)                   # list everything in the directory
                    target = [i for i in folder if ".xlsx" in i]    # find the excel file
                    data_file = new_path + '/' + target[0]          # add it to the path

                    print("current file:", data_file)

                    wb = load_workbook(data_file)
                    ws = wb['Sheet1']

                    rows = list(ws.rows)
                    cols = list(ws.columns)

                    # remove spaces in column names to make it work for SQL
                    for i in range(len(rows[0])):
                        if rows[0][i].value.find(" ") != -1:
                            rows[0][i].value = "_".join(rows[0][i].value.split())

                    # look for a match in the excel file
                    for i in range (1, len(rows)):
                        chrom_val = rows[i][0].value
                        pos_val = str(rows[i][1].value) # this gets parsed as an int so needs typecasting
                        ref_val = rows[i][2].value
                        alt_val = rows[i][3].value

                        # if a match is found
                        if chrom_val == extracted_info["CHROM"]  \
                            and pos_val == extracted_info["POS"] \
                            and ref_val == extracted_info["REF"] \
                            and alt_val == extracted_info["ALT"]:
                                
                            # prepare query data for db
                            query_data_variant = {
                                "variant_id": None,
                                "VAR_STRING": None,
                                "CHROM": None,
                                "POS": None,
                                "REF": None,
                                "ALT": None,
                                "GENE": None,
                                "ACMG": None,
                                "FEATURE_ID": None,
                                "EFFECT": None,
                                "HGVS_C": None,
                                "HGVS_P": None,
                                "ClinVar": None,
                                "ClinVarCONF": None,
                                "Varsome_link": None,
                                "Franklin_link": None
                            }

                            query_data_newsam = {
                                "sample_id": None,
                                "file_name": file_name,
                                "VAF": None,
                                "GT": None,
                                "DP": None,
                                "RELIABLE": None
                            }

                            query_data_instance = {
                                "variant_id": None,
                                "sample_id": None
                            }
                            # populate the query_data structure
                            j = 0
                            variant_string = ""
                            for cell in rows[i]:
                                current = rows[0][j].value # the column name
                                
                                if cell.value == '.':
                                    cell.value = None


                                if current == "VAF":
                                    query_data_newsam[current] = cell.value
                                elif current == "DP":
                                    query_data_newsam[current] = cell.value
                                    query_data_newsam["RELIABLE"] = True if cell.value != None and int(cell.value) > DP_THRESHOLD else False
                                elif current == "GT":
                                    if cell.value == "het":
                                        query_data_newsam[current] = 1
                                    elif cell.value == "hom":
                                        query_data_newsam[current] = 2

                                else:
                                    # print(f"{current}: {cell.value}}")
                                    query_data_variant[current] = cell.value
                                    if (current == "CHROM" or current == "REF"):
                                        variant_string += str(cell.value)
                                    elif (current == "POS" or current == "ALT"):
                                        variant_string += str(cell.value)

                                query_data_variant["VAR_STRING"] = query_info["VAR_STRING"]
                                j += 1

                            cursor.execute("SET @var_id = uuid()")
                            cursor.execute("SET @sam_id = uuid()")

                            query_variant = (
                                    "INSERT INTO variant "
                                    "(variant_id, VAR_STRING, CHROM, POS, REF, ALT, GENE, ACMG, FEATURE_ID, "
                                    "EFFECT, HGVS_C, HGVS_P, ClinVar, ClinVarCONF, Varsome_link, Franklin_link) "
                                    "VALUES (@var_id, %(VAR_STRING)s, %(CHROM)s, %(POS)s, %(REF)s, %(ALT)s, %(GENE)s,"
                                    "%(ACMG)s, %(FEATURE_ID)s, %(EFFECT)s, %(HGVS_C)s, %(HGVS_P)s, %(ClinVar)s, "
                                    " %(ClinVarCONF)s, %(Varsome_link)s, %(Franklin_link)s)"
                                    )
                            cursor.execute(query_variant, query_data_variant)

                            query_sample = (
                                    "INSERT INTO sample "
                                    "(sample_id, file_name, VAF, GT, DP, RELIABLE) VALUES "
                                    "(@sam_id, %(file_name)s, %(VAF)s, %(GT)s, %(DP)s, %(RELIABLE)s)"
                                    )
                            cursor.execute(query_sample, query_data_newsam)

                            query_instance = (
                                    "INSERT INTO instance(variant_id, sample_id) VALUES (@var_id, @sam_id)"
                                    )
                            cursor.execute(query_instance, query_data_instance)
                            
                            connection.commit()
                            
                # if variant is found, check sample
                else:
                    db_file_name = [file_name.split('.')[0] + "_ann"]
                    query_sample_check = ("SELECT s.sample_id FROM sample s JOIN instance i ON s.sample_id = i.sample_id "
                                    "JOIN variant v ON v.variant_id = i.variant_id WHERE v.VAR_STRING = %(VAR_STRING)s"
                                    "AND s.file_name = %(db_file_name)s")
                    
                    query_sample_check_data = {"db_file_name": db_file_name[0], "VAR_STRING": query_info["VAR_STRING"]}
                    cursor.execute(query_sample_check, query_sample_check_data)
                    sam_res = cursor.fetchall()

                    print("samres:", sam_res)
                    
                    # if sample is also present, append to tsv
                    if sam_res != []:
                        print(f"variant {query_info['VAR_STRING']} and sample {db_file_name[0]} already present, writing variant into tsv")
                        
                        formatted_res = [list(i) for i in res] # res returns a list of tuples by default, needs to be list of lists
                        target_vals = [formatted_res[0][2],formatted_res[0][3],formatted_res[0][4],formatted_res[0][5]]
                        tsv_writer.writerow(target_vals)
                    elif sam_res == []:
                        print(f"variant {query_info['VAR_STRING']} present, missing sample {db_file_name[0]}. Creating new instance of sample in db")

                        # get varstring id
                        query_vs_id = ("SELECT variant_id FROM variant WHERE VAR_STRING = %(VAR_STRING)s")
                        cursor.execute(query_vs_id, query_info)
                        vs_id = cursor.fetchone()

                        print("acquired id =",vs_id)

                        # create new sample in db
                        new_path = file_path.replace(file_name, '')     # strip path of the vcf file
                        folder = os.listdir(new_path)                   # list everything in the directory
                        target = [i for i in folder if ".xlsx" in i]    # find the excel file
                        data_file = new_path + '/' + target[0]          # add it to the path
                        wb = load_workbook(data_file)
                        ws = wb['Sheet1']
                        rows = list(ws.rows)

                        # remove spaces in column names to make it work for SQL
                        for i in range(len(rows[0])):
                            if rows[0][i].value.find(" ") != -1:
                                rows[0][i].value = "_".join(rows[0][i].value.split())

                        # look for a match in the excel file
                        for i in range (1, len(rows)):
                            chrom_val = rows[i][0].value
                            pos_val = str(rows[i][1].value) # this gets parsed as int so needs typecasting
                            ref_val = rows[i][2].value
                            alt_val = rows[i][3].value

                            # if a matching varstring is found
                            if chrom_val == extracted_info["CHROM"]  \
                                and pos_val == extracted_info["POS"] \
                                and ref_val == extracted_info["REF"] \
                                and alt_val == extracted_info["ALT"]:

                                query_data_newsam = {
                                "sample_id": None,
                                "file_name": db_file_name[0],
                                "VAF": None,
                                "GT": None,
                                "DP": None,
                                "RELIABLE": None
                                }

                                j = 0
                                variant_string = ""
                                for cell in rows[i]:
                                    current = rows[0][j].value # the column name
                                    
                                    if cell.value == '.':
                                        cell.value = None

                                    if current == "VAF":
                                        query_data_newsam[current] = cell.value
                                    elif current == "DP":
                                        query_data_newsam[current] = cell.value
                                        query_data_newsam["RELIABLE"] = True if cell.value != None and int(cell.value) > DP_THRESHOLD else False
                                    elif current == "GT":
                                        if cell.value == "het":
                                            query_data_newsam[current] = 1
                                        elif cell.value == "hom":
                                            query_data_newsam[current] = 2
                                    j += 1

                        cursor.execute("SET @sam_id = uuid()")

                        query_sample = (
                                    "INSERT INTO sample "
                                    "(sample_id, file_name, VAF, GT, DP, RELIABLE) VALUES "
                                    "(@sam_id, %(file_name)s, %(VAF)s, %(GT)s, %(DP)s, %(RELIABLE)s)"
                                    )
                        cursor.execute(query_sample, query_data_newsam)

                        # create new instance in db
                        query_instance = (
                                    "INSERT INTO instance(variant_id, sample_id) VALUES (%s, @sam_id)"
                                    )
                        cursor.execute(query_instance, vs_id)
                        print("added new sample to db")
                        connection.commit()

    except Error as e:
        print("Error connecting to database,", e)

    finally:
        if connection:
            if connection.is_connected():
                cursor.close()
                connection.close()
                print("Connection closed")


def main():
    # path = "../esempio_dati/CARDIOPRO-CQ1-AA33/CARDIOPRO-CQ1-AA33.vcf"
    path = "../esempio_dati/CARDIOPRO-CQ1-HD793/CARDIOPRO-CQ1-HD793.vcf"
    # path = "../esempio_dati/test_sample/test_sample.vcf"

    vcf_scraper(path)
main()